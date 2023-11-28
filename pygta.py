#!/usr/local/bin/python3

from openpyxl import Workbook
import openpyxl.utils as opxu
#import time
import re
import argparse as ap
import os
import datetime as dt

'''
sheet with highest level accummulate
'''

debug = False
reknr={}
reknr['huib'] = 'NL21INGB0004223216'
reknr['lies'] = 'NL34INGB0004261394'
reknr['enof'] = 'NL33INGB0004835652'
#print reknr.keys()

#still write a wie column

codecols=('code','naar','match','post','subp','tag')
gtcols=('date','naam','from','naar','code','afbij','val','soort','desc')
outcols=('date','wie','post','subp','muta','tag','from',
                'naar','code','afbij','val','soort','naam','desc')


def GetArgs():
    '''
    Get the arguments to reading the data
    '''
    parser = ap.ArgumentParser(description='Split gt csv transactions')

    parser.add_argument('-a','--all', action='store_true',
                        help='single year or find all')
    parser.add_argument('-o','--root',
                        default = 'tmp',
                        help='Root for output')
    parser.add_argument('-g','--gtdata',
                        default = 'testin.csv',
                        help='ING data file input')
    parser.add_argument('-c','--codes',
                        default = 'pygtacodes4.csv',
                        help='CSV file with codes input')
    parser.add_argument('-p','--projection', action='store_true',
                        help='do end of year projection')
    args = parser.parse_args()
    return args

def readCodeFile(filename):
    #this is the code file

    def readcodecsv(res):
        '''
        Reads codes, we store code, some match key word and tegenrekening
        '''
        iline = 0
        result = []
        for line in res:
            iline = iline + 1
            #print line
            line = line.strip('\n')
            tmp = line.split(',')
            cols = {}
            for i in range(len(codecols)):
                cols[codecols[i]] = tmp[i]
            result.append(cols)
            #print result            #list of match items
        return result

    codedir = './'

    #print(os.listdir())
    #afile = open( codedir+filename, 'rU')
    afile = open( codedir+filename, 'r')
    code = readcodecsv(afile)
    print('Read codes        from {:16s} found {:5d} records'.format(filename,len(code)))
    afile.close()
    return code

def readGtFile(filename):
    '''
    find the gt file
    '''
    def readgtcsv(res):
        '''
        Reads gt data, which are all in strings and has , for decimal point
        velden zijn datum, naam, rekening, tegenrekening, code, afbij, bedrag, mutatiesoort, mededeling
        '''
        iline = 0
        result = []
        for line in res:
            iline = iline + 1
            line = line.strip('\n')
            tmp = line.split('","')
            tmp[0] = tmp[0].lstrip('"')
            tmp[-1] = tmp[-1].rstrip('"') 
            if (iline > 1):
                cols = {}
                for i in range(len(gtcols)):
                    if (gtcols[i] == 'val'):
                        cols[gtcols[i]] = float(tmp[i].replace(',','.'))
                    else:
                        cols[gtcols[i]] = tmp[i]
                cols['muta'] = cols['val']
                if (cols['afbij'] == 'Af'): cols['muta'] *= -1
                cols['wie'] = '???'
                for wie in list(reknr.keys()):
                    if (reknr[wie] == cols['from']): cols['wie']=wie
                result.append( cols )
        return result

    activdir = './'
    afile = open( activdir+filename, 'r')
    activ = readgtcsv(afile)
    print('')
    print('Read transactions from {:16s} found {:5d} records'.format(filename,len(activ)))
    afile.close()
    return activ

def splitReks( data ):
    '''
    simple counting, just for 3 names and af en bij
    '''
    result = {}
    for name in list(reknr.keys()):
        result[name] = {}
        result[name]['n'] = 0
        result[name]['af'] = 0
        result[name]['bij'] = 0
    resolved = False
    for trans in data:
        for name in list(reknr.keys()):
            if (trans['from'] == reknr[name]):
                result[name]['n'] = result[name]['n']  + 1
                muta = trans['afbij'].lower()
                result[name][muta] = result[name][muta] + trans['val']
                resolved = True
        if (not resolved):
            print('mystery:',trans['from'])
    return result

def compareRekno( post, trans):
    ''' retrun True if post and trans rek numbers match.
    This is complicated by the fact that they can be valid IBAN or old fashioned digits only
    And sometimes weird strings occur'''
    match = False
    if (post=='' or trans ==''): return
    
    reiban = re.compile(r'^NL\d{2}\S{4}(\d{10})$')
    renumb = re.compile(r'^\d+$')
    #iban numbers or account numbers
    reibantrans = reiban.match(trans)
    renumbtrans = renumb.match(trans)
    reibanposts = reiban.match(post)
    renumbposts = renumb.match(post)

    try:
        if (reibantrans):
            shtrans = int(reibantrans.group(1))
        elif (renumbtrans):
            shtrans = int(trans)
        else:
            shtrans = -1
        if (reibanposts):
            shpost = int(reibanposts.group(1))
        elif (renumbposts):
            shpost = int(post)
        else:
            shpost = -2
        if  (shpost == shtrans):
            match = True
    except:
        print('Not recognized ',post,' cf ',trans) 
        
    #if (post.count('2712') and trans.count('2712')): print post,trans,shpost,shtrans,match   
    return match

def compareNaamOrDesc(naam,desc,keyw):
    match = False
    if (keyw == ''):
        match = False
    else:
        if (naam.upper().count(keyw.upper()) > 0): match = True
        if (desc.upper().count(keyw.upper()) > 0): match = True
    return match

def comparePostTrans(post, trans):
    match = False
    #print 'debug',trans['naam'],post['match'],trans['naam'].count(post['match'])
    if (post['code'] == trans['code']):
        #codes (e.g. BA for betaalautomaat) match
        #first mode is when posts are blank
        if ( post['naar'] == '' and post['match'] == ''):
            #print 'match? 1',trans, post
            match = True
        #2nd if the reks match as well            
        elif ( compareRekno(post['naar'],trans['naar']) and post['match'] == '' ):
            #print 'match? 2',trans, post
            match = True
        #of if the string is found in either and the post[naar] is blank
        elif (compareNaamOrDesc(trans['naam'],trans['desc'],post['match']) 
            and post['naar']=='' ):
            #A string can match in the name or desc, where it is always upper
            #print 'match? 3',trans, post
            match = True
    #print '\n','Match:',match
    #print 'Trans:',trans
    #print 'Posts:',post
    return match

def matchCodes( data, code ):
    '''
    This returns stats on codes being matched and a list of unmatched transactions
    '''
    unposts = []
    for trans in data:
        found = False
        for post in code:
            if (comparePostTrans(post,trans)):
                found = True
                #countPost(countp[post['tag']],trans)
                trans['tag'] = post['tag']
                trans['post'] = post['post']
                trans['subp'] = post['subp']
        if (not found):
            newp = {}
            newp['code'] = trans['code']
            newp['naar'] = trans['naar']
            newp['naam'] = trans['naam']
            newp['desc'] = trans['desc']
            newp['post'] = ''
            newp['subp'] = ''
            trans['tag'] = 'niet'
            trans['post'] = 'niet'
            trans['subp'] = 'niet'
            unposts.append(newp)

    return unposts

def exportUnPost( posten, filename ):
    #this writes the new posts to file
    res = open(filename, 'w')
    for post in posten:
        res.write('{},{},{},{},{}\n'.format(post['code'],post['naar'],
                                         post['naam'],post['desc'],
                                         post['post'], post['subp']) )
    res.close()
    return

def reportMatchStats(data) :
    #report on the number and accumulated amount of matched mutations

    totaf = 0; unaf = 0
    unbij = 0; totbij = 0
    ntot = 0; nun = 0
    for mut in data:
        ntot += 1
        if (mut['afbij']=='Af'):
            totaf += mut['val']
        else:
            totbij += mut['val']
        if (mut['tag'] == 'niet'):
            nun += 1
            if (mut['afbij']=='Af'):
                unaf += mut['val']
            else:
                unbij += mut['val']

    print('Number of mutation unassigned: {:5d}/{:5d}'.format(nun, ntot))
    print('Fraction of mutations unassigned -,+: {:5.1f},{:5.1f}'.format(100*unaf/totaf, 100*unbij/totbij))

    return {'unaf': unaf, 'totaf': totaf, 'unbij': unbij, 'totbij': totbij}

def reportbyRek( data ):
    '''
    Splits that data by account number and sums the mutations
    '''
    #print year[42]
    totaf = 0
    totbij = 0
    result = splitReks(data)
    for rek in list(result.keys()):
        print('  - {:4s} -,+ {:9.2f},{:9.2f} ntrans: {:5d}'.format(rek,result[rek]['af'],result[rek]['bij'],result[rek]['n']))
        totaf += result[rek]['af']
        totbij += result[rek]['bij']
    return

def printProg( prog ):
    sum={}
    for pp in list(prog[0].keys()):
        sum[pp] = 0
    for year in prog:
        for pp in list(year.keys()):
            sum[pp] += year[pp]
    print('TOTAL Fraction of mutations unassigned -,+: {:5.1f},{:5.1f}'.format(100*sum['unaf']/sum['totaf'], 
                                                                        100*sum['unbij']/sum['totbij']))
    print('       --,++: {:10.2f}/{:10.2f} , {:10.2f}/{:10.2f}'.format(sum['unaf'],sum['totaf'], 
                                                                       sum['unbij'],sum['totbij']))

def iMonth(field):
    redate = re.compile(r'^\d{4}(\d{2})\d{2}$')
    mon = redate.match(field)
    #print mon.group(1)
    return int(mon.group(1))

def accumPosts( data, posts ):

    budget = []
    for im in range(12):
        budget.append( {} )
        for wie in list(reknr.keys()):
            budget[im][wie] = {}
            for post in posts:
                bin = {};
                bin['n'] = 0;bin['af'] = 0;bin['bij'] = 0
                #print post, bin
                budget[im][wie][post[0]]=bin
    for trans in data:
    #not yet updated to include subposts
        post = trans['post']
        wie = trans['wie']
        im = iMonth(trans['date'])-1
        budget[im][wie][post]['n'] += 1
        if (trans['afbij'] == 'Af'):
            budget[im][wie][post]['af'] += trans['val']
        elif (trans['afbij'] == 'Bij'):
            budget[im][wie][post]['bij'] += trans['val']
        else:
            print('Nee toch?!')
    return budget

def sortPosts( codes ):
    #return a list of post, subp tuples, prioritised by post
    prioposts = ['inkomsten','vastelasten','transacties','auto','woning',
    			'gezondheid','educatie','huishouden','kleding','mobiliteit',
    			'abonnementen','goeddoel','vermaak','vakantie','niet']
    tmpposts = []
    for item in codes:
        if ((item['post'],item['subp']) not in tmpposts):
            tmpposts.append((item['post'],item['subp']))
    posts = []
    #the prio items go first
    for prio in prioposts:
        for i in range(len(tmpposts)):
            if (prio == tmpposts[i][0]):
                posts.append(tmpposts[i])
    for item in tmpposts:
        Found = False
        for i in range(len(posts)):
           if (item == posts[i]):
               Found = True
        if not Found: posts.append(item)
    posts.append(('niet','niet'))
    #check the subposts are unique
    checksubp = []
    for item in posts:
        if item[1] in checksubp:
            message = 'sub posts should be unique: '+str(item)
            raise Exception(message)
        else:
            checksubp.append(item[1])
    
    #print posts
    return posts

def auxFiles( root ):
    fileunset = root+'_unset.csv'
    filebudg = root+'_budg.xlsx'
    return (fileunset, filebudg)

def findInputs():
    pathnow = '.'
    refile = re.compile(r'^Alle_rekeningen_\d{2}-\d{2}-(\S+)_\d{2}-\d{2}-(\S+).csv$')
    filename = []
    found = False
    filelist = os.listdir(pathnow)
    for file in filelist:
        yesdata = refile.match(file)
        if yesdata:
            filename.append(file)
    return filename

def writeExcel(posts,data,file,do_x):
    #writes excel sheets with 3 sort of sheets
    def fillCell_Sumif(ws,ix,iy,sheet,colfind,colkey,colrow,colsum,corr):
        formula='=SUMIF({}!{},{}{},{}!{})/{}'.format(sheet,colfind,colkey,str(colrow),sheet,colsum,corr)     
        #print ix,iy,formula
        ws.cell(column=ix,row=iy,value=formula)

    def fillCell_3Sumif(ws,ix,iy,sheets,colfind,colkey,colrow,colsum,corr):
        formula='='
        for sheet in sheets:
            formula+='SUMIF({}!{},{}{},{}!{})/{}'.format(sheet,colfind,colkey,str(colrow),sheet,colsum,corr)
            formula+='+'
        formula = formula[0:-1]
        #print ix,iy,formula
        ws.cell(column=ix,row=iy,value=formula)
    
    def checkYears(data):
        #estimate what the completeness fraction for each year is, for all years
        def gtdate(gtstr):
            year = int(gtstr[0:4])
            mon = int(gtstr[4:6])
            day = int(gtstr[6:8])
            #print year,mon,day
            return dt.datetime(year,mon,day)

        fullyears = {}
        for year in data:
            oldest = '20991231'
            newest = '19001127'
            for mut in year:
                if (mut['date'] > newest): newest = mut['date']
                if (mut['date'] < oldest): oldest = mut['date']
            #print 'Checking:',oldest,newest
            olddate = gtdate(oldest)
            newdate = gtdate(newest)
            yearfrac = (newdate.toordinal()-olddate.toordinal())/365.
            #print yearfrac,olddate.year
            if (yearfrac > 0.98 and yearfrac < 1.02): 
                yearfrac = 1.0
            else:
                print('Year ',olddate.year,' not complete ',yearfrac)
            fullyears[str(olddate.year)]=yearfrac
            
        return fullyears
                
    
    def writeYears(data):
        #sheets are just the names
        sheets = [] 
        #write each year, and split by account    
        for year in reversed(data):
            for wie in list(reknr.keys()):
                shname = year[1]['date'][0:4]+wie
                sheets.append(shname)
                ws2 = wb.create_sheet()
                ws2.title = shname
                ws2.append(outcols)
                for mut in year:
                    if (mut['wie'] == wie):
                        tmp = []
                        for post in outcols:
                            tmp.append(mut[post])
                        ws2.append(tmp)
        return sheets

    def writeMatrix(posts, sheets, fractions):
        #now write posts
        ws1 = wb.active
        ws1.title = 'matrix'
        cols = ['post','subp']
        cols += sheets
        ws1.append(cols) 
        for post in posts: 
            ws1.append(post)
        #now fill the rest with sumifs
        ix=2
        colsub = 'D:D';colmut='E:E';colkey='B'
        for rek in sheets:
            corr = 1
            for fracy in list(fractions.keys()):
                if ( fracy == rek[0:4]):
                    #print 'check',fracy,fractions[fracy]
                    corr = fractions[fracy]
            ix += 1
            iy = 1
            for cat in posts:
    	        iy +=1
    	        fillCell_Sumif(ws1,ix,iy,rek,colsub,colkey,iy,colmut,corr)
                                
    def writeSum(posts, sheets, fractions):
        #now write posts
        ws0 = wb.create_sheet(index=0)
        ws0.title = 'sum'
        cols = ['post']
        caly = []
        pyear='00'
        sumy = []
        thisy= []
        for item in sheets:
            if (item[0:4] != pyear):
                pyear = item[0:4]
                caly.append(pyear)
                if (len(thisy)>0):
                    sumy.append(thisy)
                thisy = [item]
            else:
                thisy.append(item)
        if (len(thisy)>0):
                sumy.append(thisy)
        mposts = []
        for item in posts:
            if (not item[0] in mposts):
                mposts.append(item[0])
        #print mposts
        #print caly
        #print sumy
        cols += caly
        
        ws0.append(cols) 
        for post in mposts: 
            ws0.append([post])
        #now fill the rest with sumifs
        ix=1
        colsub = 'C:C';colmut='E:E';colkey='A'
        for icol in range(len(caly)):
            #print 'TEST',caly[icol],fractions[caly[icol]]
            corr = fractions[caly[icol]]
            ix += 1
            iy = 1
            for cat in mposts:
                iy +=1
                fillCell_3Sumif(ws0,ix,iy,sumy[icol],colsub,colkey,iy,colmut,corr)

    #assumes data is list of dictionaries with same keys.
    wb = Workbook()
    sheets = writeYears(data)
    annfractions = checkYears(data)
    print('fracs: ', annfractions)
    
    if (not do_x):
        print('giving cumulatives no projections')
        for fracy in list(annfractions.keys()):
            annfractions[fracy]=1.0
    print('fracs again:', annfractions)
    #huibhier, must divide projection, but trick is to also write originals
    
    writeMatrix(posts, sheets, annfractions)
    writeSum(posts,sheets, annfractions)

    #Write a top sheet
    #ws2 = wb.create_sheet(0)
    #ws2.title = 'top'
    #cols = ['posts']
    #we need to sum 3 sheets every time
    #create a list of dictionary
    
    wb.save(file)

#-----------------------------------------------------------------------------------
#  MAIN

opts = GetArgs()
(fileunset, filebudg) = auxFiles( opts.root )

if (opts.all):
    print('Mode is All')
else:
    print('Mode is Single')
print(('Digesting {} using {}.'.format(opts.gtdata,opts.codes) ))
print(('Budget in {} unset Mutations in {}.'.format(filebudg,fileunset) ))

codes = readCodeFile(opts.codes)
posts = sortPosts(codes) 

files = []
unmatched = []
budget = []
progress = []

if (opts.all):
    files += findInputs()
    print(files)
else:
    files.append(opts.gtdata)
    
for iy in range(len(files)):
    year = readGtFile(files[iy])

    #print year
    unmatched += matchCodes(year, codes)

    reportbyRek(year)

    progress.append( reportMatchStats(year ))

    budget.append( year )

    #exportBudget(budget, filebudg)

    #exportTrans( year, opts.root, yearstr[iy] )

exportUnPost(unmatched, fileunset )
#must write posts
writeExcel(posts,budget,filebudg,opts.projection)

printProg( progress )



